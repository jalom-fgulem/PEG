"""
Módulo de Ingresos SGPEG — generador SUENLACE.DAT para facturas/cobros emitidos.

Porta la lógica de suenlace/src/App.jsx al backend Python.
Áreas: hospital, colegio, mule, podologia.
Modos: factura_con_iva, factura_sin_iva, cobro_simple.
"""

import re
import unicodedata
from datetime import datetime, date, timedelta
from typing import Optional
import openpyxl

# ── Mock storage ──────────────────────────────────────────────────────────────

LOTES_INGRESOS: list[dict] = []
REGISTROS_INGRESOS: list[dict] = []
_lote_id_seq = 0
_reg_id_seq = 0


def _next_lote_id() -> int:
    global _lote_id_seq
    _lote_id_seq += 1
    return _lote_id_seq


def _next_reg_id() -> int:
    global _reg_id_seq
    _reg_id_seq += 1
    return _reg_id_seq


def listar_lotes(area: Optional[str] = None, estado: Optional[str] = None) -> list[dict]:
    items = list(LOTES_INGRESOS)
    if area:
        items = [l for l in items if l["area"] == area]
    if estado:
        items = [l for l in items if l["estado"] == estado]
    return sorted(items, key=lambda l: l["fecha_importacion"], reverse=True)


def obtener_lote(id_lote: int) -> Optional[dict]:
    return next((l for l in LOTES_INGRESOS if l["id_lote"] == id_lote), None)


def obtener_registros_lote(id_lote: int) -> list[dict]:
    return [r for r in REGISTROS_INGRESOS if r["id_lote"] == id_lote]


def marcar_exportado(id_lote: int):
    lote = obtener_lote(id_lote)
    if lote:
        lote["estado"] = "EXPORTADO"


# ── CONFIG — espejo del CONFIG del React ──────────────────────────────────────

HOSPITAL_COBRO_FIELDS = {
    "numero":  ["Factura"],
    "cliente": ["Cliente"],
    "fecha":   ["Fecha cobro"],
    "importe": ["Cobrado"],
}

MULE_FIELDS = {
    "numero":  ["Número de Factura"],
    "cliente": ["Cliente"],
    "fecha":   ["Fecha"],
    "importe": ["Total", "Importe"],
}

CPOD_FACTURA_FIELDS = {
    "numero":  ["N. Factura"],
    "cliente": ["Cliente"],
    "nif":     ["DNI/CIF Cliente"],
    "fecha":   ["Fecha"],
    "importe": ["Total", "Importe"],
}

CPOD_COBRO_FIELDS = {
    "numero":  ["Factura"],
    "cliente": ["Cliente"],
    "fecha":   ["Fecha cobro"],
    "importe": ["Cobrado"],
}

CONFIG = {
    "hospital": {
        "facturas": {
            "implemented": True,
            "mode": "factura_con_iva",
            "descripcionCliente": "Facturas HV",
            "descripcionIngreso": "Facturas HV",
            "cuentaCliente":   "4400001",
            "cuentaIngresos":  "7590005",
            "cuentaIVA":       "4770000",
            "cuentaRetencion": "4730000",
            "anal1": "6",
            "anal2": "1",
            "ivaFieldMode": "excel_total_is_vat",
            "outputFilename": "FHV.dat",
            "fields": {
                "numero":   ["Número de Factura", "Nº Factura", "FACTURA"],
                "cliente":  ["Cliente", "CLIENTE"],
                "nif":      ["DNI", "NIF"],
                "fecha":    ["Fecha", "FECHA"],
                "base":     ["Base imponible", "BASE IMPONIBLE"],
                "ivaPct":   ["Tipo Impuesto", "TIPO IMPUESTO", "IVA"],
                "ivaCuota": ["Total", "CUOTA IVA", "IMPORTE IVA"],
            },
        },
        "cobros": {
            "implemented": True,
            "mode": "cobro_simple",
            "descripcionCliente": "Cobros HV",
            "descripcionBanco":   "Cobros HV",
            "cuentaCliente": "4400001",
            "cuentaBanco":   "5720104",
            "outputFilename": "CHV.dat",
            "fields": HOSPITAL_COBRO_FIELDS,
            "specialSheetName": "Informe de Cobros (2)",
        },
        "pagos": {"implemented": False},
    },
    "colegio": {
        "facturas": {
            "implemented": True,
            "mode": "factura_con_iva",
            "descripcionCliente": "Facturas CM",
            "descripcionIngreso": "Facturas CM",
            "cuentaCliente":   "4400203",
            "cuentaIngresos":  "7590007",
            "cuentaIVA":       "4770000",
            "cuentaRetencion": "4730000",
            "anal1": "2",
            "anal2": "5",
            "ivaFieldMode": "excel_split_base_vat_total",
            "fixedIvaPct": 10.0,
            "outputFilename": "FCM.dat",
            "fields": {
                "numero":   ["Nº DOC", "NUM DOC", "FACTURA"],
                "cliente":  ["RESIDENTE", "Cliente", "CLIENTE"],
                "nif":      ["DNI", "NIF"],
                "fecha":    ["FECHA EMISIÓN", "Fecha Emisión", "FECHA"],
                "base":     ["IMPORTE BASE (€)", "BASE IMPONIBLE"],
                "ivaCuota": ["IMPORTE IVA (€)", "CUOTA IVA"],
                "total":    ["IMPORTE (€)", "TOTAL"],
            },
        },
        "cobros": {
            "implemented": True,
            "mode": "cobro_simple",
            "descripcionCliente": "Cobros CM",
            "descripcionBanco":   "Cobros CM",
            "cuentaCliente": "4400203",
            "cuentaBanco":   "5720104",
            "outputFilename": "CCM.dat",
            "fields": {
                "numero":  ["Nº DOC"],
                "cliente": ["RESIDENTE"],
                "fecha":   ["FECHA VENCIMIENTO"],
                "importe": ["IMPORTE (€)"],
            },
        },
        "pagos": {"implemented": False},
    },
    "mule": {
        "facturas": {
            "implemented": True,
            "mode": "factura_sin_iva",
            "descripcionCliente": "Facturas MULE",
            "descripcionIngreso": "Facturas MULE",
            "cuentaCliente":  "4400204",
            "cuentaIngresos": "7590001",
            "anal1": "2",
            "anal2": "2",
            "outputFilename": "FMULE.dat",
            "fields": MULE_FIELDS,
        },
        "cobros": {
            "implemented": True,
            "mode": "cobro_simple",
            "descripcionCliente": "Cobros MULE",
            "descripcionBanco":   "Cobros MULE",
            "cuentaCliente": "4400204",
            "cuentaBanco":   "5720104",
            "outputFilename": "CMULE.dat",
            "fields": {
                "numero":  ["Número de Factura"],
                "cliente": ["Cliente"],
                "fecha":   ["Fecha"],
                "importe": ["Total", "Importe"],
            },
        },
        "pagos": {"implemented": False},
    },
    "podologia": {
        "facturas": {
            "implemented": True,
            "mode": "factura_sin_iva",
            "descripcionCliente": "Facturas CPOD",
            "descripcionIngreso": "Facturas CPOD",
            "cuentaCliente":  "4400002",
            "cuentaIngresos": "7590006",
            "anal1": "6",
            "anal2": "2",
            "outputFilename": "FCPOD.dat",
            "fields": CPOD_FACTURA_FIELDS,
        },
        "cobros": {
            "implemented": True,
            "mode": "cobro_simple",
            "descripcionCliente": "Cobros CPOD",
            "descripcionBanco":   "Cobros CPOD",
            "cuentaCliente": "4400002",
            "cuentaBanco":   "5720104",
            "outputFilename": "CCPOD.dat",
            "fields": CPOD_COBRO_FIELDS,
        },
        "pagos": {"implemented": False},
    },
    "fgulem": {
        "facturas": {"implemented": False},
        "cobros":   {"implemented": False},
        "pagos":    {"implemented": False},
    },
}

LABELS_AREA = {
    "hospital":  "Hospital Veterinario",
    "colegio":   "Colegio Mayor",
    "mule":      "MULE",
    "podologia": "Clínica de Podología",
    "fgulem":    "FGULEM",
}

LABELS_PROCESO = {
    "facturas": "Facturas",
    "cobros":   "Cobros",
    "pagos":    "Pagos",
}

A3_CODIGOS = {
    "pruebas": "00006",
    "oficial": "00005",
}

# ── Helpers de texto ──────────────────────────────────────────────────────────

def sanitize_text(value) -> str:
    s = str(value) if value is not None else ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("ñ", "n").replace("Ñ", "N")
    s = re.sub(r"[^A-Za-z0-9 .,_\-\/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def pad_right(value, length: int) -> str:
    return str(value or "")[:length].ljust(length)


def normalize_account(value) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[:12].ljust(12, "0")


def normalize_reference(value) -> str:
    s = sanitize_text(value)
    s = re.sub(r"[/\-]", "", s)
    s = re.sub(r"\s+", "", s)
    return s[:10]


def parse_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    cleaned = text.replace("€", "").replace(" ", "")
    cleaned = re.sub(r"\.(?=\d{3}(\D|$))", "", cleaned)
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def infer_iva_percent(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(value))
    if m:
        return float(m.group(0).replace(",", "."))
    return 0.0


_EXCEL_EPOCH = datetime(1899, 12, 30)


def format_date_to_a3(value) -> str:
    """Convierte varios formatos de fecha a aaaammdd."""
    if value is None or value == "":
        return ""

    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.strftime("%Y%m%d")
        return value.strftime("%Y%m%d")

    if isinstance(value, (int, float)):
        try:
            dt = _EXCEL_EPOCH + timedelta(days=int(value))
            return dt.strftime("%Y%m%d")
        except Exception:
            return ""

    text = str(value).strip()

    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", text)
    if m:
        return f"{m.group(1)}{m.group(2).zfill(2)}{m.group(3).zfill(2)}"

    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})$", text)
    if m:
        year = ("20" + m.group(3)) if len(m.group(3)) == 2 else m.group(3)
        return f"{year}{m.group(2).zfill(2)}{m.group(1).zfill(2)}"

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.strftime("%Y%m%d")
        except ValueError:
            pass

    return ""


def format_signed_amount(value) -> str:
    """Formatea importe a +NNNNNNNNNN.DD (14 chars)."""
    n = float(value or 0)
    sign = "-" if n < 0 else "+"
    abs_val = abs(n)
    int_part = int(abs_val)
    dec_part = round((abs_val - int_part) * 100)
    return f"{sign}{str(int_part).zfill(10)}.{str(dec_part).zfill(2)}"


def pick_field(row: dict, candidates: list) -> object:
    """Busca el valor de la primera columna que coincida (case-insensitive) con algún candidato."""
    row_lower = {k.strip().lower(): v for k, v in row.items()}
    for candidate in candidates:
        val = row_lower.get(candidate.strip().lower())
        if val is not None:
            return val
    return None


# ── Constructores de registro SUENLACE (ingresos) ─────────────────────────────

def _base_record(company_code: str, fecha: str, tipo: str) -> list:
    """Devuelve array mutable de 254 espacios con campos comunes ya rellenos."""
    arr = [" "] * 254
    arr[0] = "4"
    for i, c in enumerate(str(company_code or "")[:5]):
        arr[1 + i] = c
    for i, c in enumerate(str(fecha or "")[:8]):
        arr[6 + i] = c
    arr[14] = tipo
    arr[252] = "E"
    arr[253] = "N"
    return arr


def _put(arr: list, start: int, text: str):
    """Escribe text en arr a partir de la posición start (base 1)."""
    idx = start - 1
    for i, c in enumerate(str(text or "")):
        if idx + i < len(arr):
            arr[idx + i] = c


def _to_line(arr: list) -> str:
    return "".join(arr) + "\r\n"


def build_record_1(company_code, fecha, cuenta_cliente, nombre, referencia, total, nif) -> str:
    arr = _base_record(company_code, fecha, "1")
    _put(arr, 16, normalize_account(cuenta_cliente))
    _put(arr, 28, pad_right(sanitize_text(nombre), 30))
    _put(arr, 58, "1")
    _put(arr, 59, pad_right(normalize_reference(referencia), 10))
    _put(arr, 69, "I")
    _put(arr, 70, pad_right(sanitize_text(nombre), 30))
    _put(arr, 100, format_signed_amount(total))
    _put(arr, 176, pad_right(sanitize_text(str(nif or ""))[:14], 14))
    _put(arr, 190, pad_right(sanitize_text(nombre), 40))
    assert len(arr) == 254
    return _to_line(arr)


def build_record_9(
    company_code, fecha, cuenta_ingresos, descripcion, referencia,
    base, iva_pct, iva_cuota, cuenta_iva, cuenta_retencion, has_analitica
) -> str:
    arr = _base_record(company_code, fecha, "9")
    _put(arr, 16, normalize_account(cuenta_ingresos))
    _put(arr, 28, pad_right(sanitize_text(descripcion), 30))
    _put(arr, 58, "C")
    _put(arr, 59, pad_right(normalize_reference(referencia), 10))
    _put(arr, 69, "U")
    _put(arr, 70, pad_right(sanitize_text(descripcion), 30))
    _put(arr, 100, "01")
    _put(arr, 102, format_signed_amount(base))
    _put(arr, 116, f"{float(iva_pct or 0):05.2f}")
    _put(arr, 121, format_signed_amount(iva_cuota))
    _put(arr, 154, "00.00")
    _put(arr, 159, format_signed_amount(0))
    _put(arr, 175, "N" if (iva_pct or 0) == 0 else "S")
    _put(arr, 192, normalize_account(cuenta_iva))
    _put(arr, 216, normalize_account(cuenta_retencion))
    _put(arr, 252, "S" if has_analitica else " ")
    assert len(arr) == 254
    return _to_line(arr)


def build_record_0(
    company_code, fecha, cuenta, descripcion_cuenta, debe_haber,
    referencia, linea, descripcion_apunte, importe, analitica
) -> str:
    arr = _base_record(company_code, fecha, "0")
    _put(arr, 16, normalize_account(cuenta))
    _put(arr, 28, pad_right(sanitize_text(descripcion_cuenta), 30))
    _put(arr, 58, debe_haber)
    _put(arr, 59, pad_right(normalize_reference(referencia), 10))
    _put(arr, 69, linea)
    _put(arr, 70, pad_right(sanitize_text(descripcion_apunte), 30))
    _put(arr, 100, format_signed_amount(importe))
    _put(arr, 252, "S" if analitica else " ")
    assert len(arr) == 254
    return _to_line(arr)


def build_record_d(
    company_code, fecha, cuenta, descripcion, importe_base, anal1, anal2, linea_apunte="002"
) -> str:
    arr = _base_record(company_code, fecha, "D")
    _put(arr, 16, normalize_account(cuenta))
    _put(arr, 28, pad_right(sanitize_text(descripcion), 30))
    _put(arr, 59, format_signed_amount(importe_base))
    _put(arr, 73, str(linea_apunte)[:3])
    _put(arr, 76, "I")
    _put(arr, 77, pad_right(str(anal1), 4))
    _put(arr, 81, pad_right(str(anal2), 4))
    _put(arr, 85, pad_right("", 4))
    _put(arr, 89, pad_right("", 4))
    _put(arr, 93,  pad_right(f"ANAL {anal1}", 30))
    _put(arr, 123, pad_right(f"ANAL {anal2}", 30))
    _put(arr, 153, pad_right("", 30))
    _put(arr, 183, pad_right("", 30))
    _put(arr, 213, format_signed_amount(importe_base))
    _put(arr, 227, "100.00")
    assert len(arr) == 254
    return _to_line(arr)


# ── Procesadores de filas ─────────────────────────────────────────────────────

def _process_factura_con_iva(rows: list[dict], cfg: dict, company_code: str) -> dict:
    processed, errors, dat = [], [], ""
    total_base = total_iva = total_general = 0.0

    for idx, row in enumerate(rows):
        try:
            numero   = pick_field(row, cfg["fields"]["numero"])
            cliente  = pick_field(row, cfg["fields"]["cliente"])
            nif      = pick_field(row, cfg["fields"].get("nif", []))
            fecha_raw = pick_field(row, cfg["fields"]["fecha"])
            base_raw  = pick_field(row, cfg["fields"]["base"])

            fecha = format_date_to_a3(fecha_raw)
            if not fecha:
                raise ValueError("Fecha inválida en factura")

            base = parse_number(base_raw)

            if not numero or not cliente or not fecha:
                raise ValueError("Faltan campos obligatorios (número, cliente, fecha)")

            if cfg["ivaFieldMode"] == "excel_total_is_vat":
                iva_cuota = parse_number(pick_field(row, cfg["fields"]["ivaCuota"]))
                iva_pct   = infer_iva_percent(pick_field(row, cfg["fields"]["ivaPct"]))
                total     = base + iva_cuota
            else:
                iva_cuota = parse_number(pick_field(row, cfg["fields"]["ivaCuota"]))
                total     = parse_number(pick_field(row, cfg["fields"]["total"]))
                iva_pct   = cfg.get("fixedIvaPct") or (
                    round((iva_cuota / base) * 100, 2) if base != 0 else 0.0
                )

            referencia = normalize_reference(numero)
            nombre     = sanitize_text(cliente)

            dat += build_record_1(
                company_code, fecha,
                cfg["cuentaCliente"], nombre, referencia, total, nif or "",
            )
            dat += build_record_9(
                company_code, fecha,
                cfg["cuentaIngresos"], cfg["descripcionIngreso"], referencia,
                base, iva_pct, iva_cuota,
                cfg["cuentaIVA"], cfg["cuentaRetencion"], has_analitica=True,
            )
            dat += build_record_d(
                company_code, fecha,
                cfg["cuentaIngresos"], cfg["descripcionIngreso"],
                base, cfg["anal1"], cfg["anal2"],
            )

            total_base    += base
            total_iva     += iva_cuota
            total_general += total

            processed.append({
                "id_registro":   None,
                "fila_excel":    idx + 2,
                "numero_factura": str(numero or ""),
                "cliente":        nombre,
                "nif":            sanitize_text(str(nif or "")),
                "fecha":          fecha,
                "base_imponible": round(base, 2),
                "porcentaje_iva": round(float(iva_pct), 2),
                "cuota_iva":      round(iva_cuota, 2),
                "total":          round(total, 2),
                "estado":         "OK",
                "error_mensaje":  "",
            })
        except Exception as exc:
            errors.append({
                "id_registro":   None,
                "fila_excel":    idx + 2,
                "numero_factura": "",
                "cliente":        "",
                "nif":            "",
                "fecha":          "",
                "base_imponible": 0,
                "porcentaje_iva": 0,
                "cuota_iva":      0,
                "total":          0,
                "estado":         "ERROR",
                "error_mensaje":  str(exc),
            })

    return {
        "dat": dat,
        "processed": processed,
        "errors": errors,
        "summary": {
            "total_base":    round(total_base, 2),
            "total_iva":     round(total_iva, 2),
            "total_general": round(total_general, 2),
        },
    }


def _process_factura_sin_iva(rows: list[dict], cfg: dict, company_code: str) -> dict:
    processed, errors, dat = [], [], ""
    total_base = 0.0

    for idx, row in enumerate(rows):
        try:
            numero    = pick_field(row, cfg["fields"]["numero"])
            cliente   = pick_field(row, cfg["fields"]["cliente"])
            fecha_raw = pick_field(row, cfg["fields"]["fecha"])
            importe_raw = pick_field(row, cfg["fields"]["importe"])

            fecha   = format_date_to_a3(fecha_raw)
            importe = parse_number(importe_raw)

            if not numero or not cliente or not fecha:
                raise ValueError("Faltan campos obligatorios (número, cliente, fecha)")

            referencia = normalize_reference(numero)
            nombre     = sanitize_text(cliente)

            dat += build_record_0(
                company_code, fecha,
                cfg["cuentaCliente"], nombre,
                "D", referencia, "I", cfg["descripcionCliente"], importe, analitica=False,
            )
            dat += build_record_0(
                company_code, fecha,
                cfg["cuentaIngresos"], cfg["descripcionIngreso"],
                "H", referencia, "U", cfg["descripcionIngreso"], importe, analitica=True,
            )
            dat += build_record_d(
                company_code, fecha,
                cfg["cuentaIngresos"], cfg["descripcionIngreso"],
                importe, cfg["anal1"], cfg["anal2"],
            )

            total_base += importe

            processed.append({
                "id_registro":   None,
                "fila_excel":    idx + 2,
                "numero_factura": str(numero or ""),
                "cliente":        nombre,
                "nif":            "",
                "fecha":          fecha,
                "base_imponible": round(importe, 2),
                "porcentaje_iva": 0,
                "cuota_iva":      0,
                "total":          round(importe, 2),
                "estado":         "OK",
                "error_mensaje":  "",
            })
        except Exception as exc:
            errors.append({
                "id_registro":   None,
                "fila_excel":    idx + 2,
                "numero_factura": "", "cliente": "", "nif": "", "fecha": "",
                "base_imponible": 0, "porcentaje_iva": 0, "cuota_iva": 0, "total": 0,
                "estado": "ERROR", "error_mensaje": str(exc),
            })

    return {
        "dat": dat,
        "processed": processed,
        "errors": errors,
        "summary": {
            "total_base":    round(total_base, 2),
            "total_iva":     0.0,
            "total_general": round(total_base, 2),
        },
    }


def _process_cobro_simple(rows: list[dict], cfg: dict, company_code: str) -> dict:
    processed, errors, dat = [], [], ""
    total_importes = 0.0

    for idx, row in enumerate(rows):
        try:
            numero    = pick_field(row, cfg["fields"]["numero"])
            cliente   = pick_field(row, cfg["fields"]["cliente"])
            fecha_raw = pick_field(row, cfg["fields"]["fecha"])
            importe_raw = pick_field(row, cfg["fields"]["importe"])

            fecha   = format_date_to_a3(fecha_raw)
            importe = parse_number(importe_raw)

            if not numero or not fecha or importe_raw is None or importe_raw == "":
                raise ValueError("Faltan campos obligatorios (número, fecha, importe)")

            referencia = normalize_reference(numero)
            nombre     = sanitize_text(str(cliente or cfg["descripcionCliente"]))

            dat += build_record_0(
                company_code, fecha,
                cfg["cuentaBanco"], "Banco",
                "D", referencia, "I", cfg["descripcionBanco"], importe, analitica=False,
            )
            dat += build_record_0(
                company_code, fecha,
                cfg["cuentaCliente"], nombre,
                "H", referencia, "U", cfg["descripcionCliente"], importe, analitica=False,
            )

            total_importes += importe

            processed.append({
                "id_registro":   None,
                "fila_excel":    idx + 2,
                "numero_factura": str(numero or ""),
                "cliente":        nombre,
                "nif":            "",
                "fecha":          fecha,
                "base_imponible": round(importe, 2),
                "porcentaje_iva": 0,
                "cuota_iva":      0,
                "total":          round(importe, 2),
                "estado":         "OK",
                "error_mensaje":  "",
            })
        except Exception as exc:
            errors.append({
                "id_registro":   None,
                "fila_excel":    idx + 2,
                "numero_factura": "", "cliente": "", "nif": "", "fecha": "",
                "base_imponible": 0, "porcentaje_iva": 0, "cuota_iva": 0, "total": 0,
                "estado": "ERROR", "error_mensaje": str(exc),
            })

    return {
        "dat": dat,
        "processed": processed,
        "errors": errors,
        "summary": {
            "total_base":    round(total_importes, 2),
            "total_iva":     0.0,
            "total_general": round(total_importes, 2),
        },
    }


# ── Lectura de Excel con openpyxl ─────────────────────────────────────────────

def _read_excel_rows(file_bytes: bytes, sheet_name: Optional[str] = None) -> tuple[list[dict], str]:
    """
    Lee un Excel y devuelve (rows, nombre_hoja_usada).
    rows es lista de dicts con {header: valor}.
    """
    wb = openpyxl.load_workbook(filename=__import__("io").BytesIO(file_bytes), data_only=True)

    if sheet_name:
        # Buscar hoja case-insensitive
        match = next(
            (n for n in wb.sheetnames if n.strip().lower() == sheet_name.strip().lower()),
            None,
        )
        if not match:
            raise ValueError(f'No se encontró la hoja "{sheet_name}" en el Excel.')
        ws = wb[match]
        used_sheet = match
    else:
        ws = wb.active
        used_sheet = ws.title

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], used_sheet

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

    rows = []
    for row in rows_iter:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})

    return rows, used_sheet


# ── Función principal de importación ─────────────────────────────────────────

def procesar_excel(
    file_bytes: bytes,
    nombre_fichero_origen: str,
    area: str,
    tipo_proceso: str,
    destino_a3: str,
    usuario_id: int,
) -> dict:
    """
    Procesa el Excel, genera el DAT, guarda lote y registros en mock.
    Devuelve el dict del lote creado (con 'dat_content' para descarga).
    """
    cfg_area = CONFIG.get(area, {})
    cfg = cfg_area.get(tipo_proceso, {})

    if not cfg.get("implemented"):
        raise ValueError(f"La combinación {area}/{tipo_proceso} no está implementada.")

    company_code = A3_CODIGOS.get(destino_a3, "00006")

    sheet_name = cfg.get("specialSheetName")
    rows, hoja_usada = _read_excel_rows(file_bytes, sheet_name)

    if not rows:
        raise ValueError("El Excel no contiene filas de datos.")

    mode = cfg["mode"]
    if mode == "factura_con_iva":
        result = _process_factura_con_iva(rows, cfg, company_code)
    elif mode == "factura_sin_iva":
        result = _process_factura_sin_iva(rows, cfg, company_code)
    elif mode == "cobro_simple":
        result = _process_cobro_simple(rows, cfg, company_code)
    else:
        raise ValueError(f"Modo de proceso desconocido: {mode}")

    summary = result["summary"]
    all_records = result["processed"] + result["errors"]
    nombre_dat = cfg.get("outputFilename", "SUENLACE.DAT")

    id_lote = _next_lote_id()
    lote = {
        "id_lote":               id_lote,
        "area":                  area,
        "tipo_proceso":          tipo_proceso,
        "destino_a3":            destino_a3,
        "fecha_importacion":     datetime.now().isoformat(),
        "usuario_id":            usuario_id,
        "total_registros":       len(all_records),
        "registros_ok":          len(result["processed"]),
        "registros_error":       len(result["errors"]),
        "importe_total_base":    summary["total_base"],
        "importe_total_iva":     summary["total_iva"],
        "importe_total":         summary["total_general"],
        "nombre_fichero_origen": nombre_fichero_origen,
        "nombre_fichero_dat":    nombre_dat,
        "estado":                "GENERADO",
        "hoja_excel":            hoja_usada,
        "filas_excel":           len(rows),
        "dat_content":           result["dat"],
    }
    LOTES_INGRESOS.append(lote)

    for rec in all_records:
        reg = dict(rec)
        reg["id_registro"] = _next_reg_id()
        reg["id_lote"]     = id_lote
        REGISTROS_INGRESOS.append(reg)

    return lote
