"""
Parser de extractos bancarios para SGPEG.

Formatos soportados:
  - AEB Norma 43 (.n43 / .43 / .sta / .csb / .txt con cabecera "11")
    · Variante estándar: D/C en pos 22 ('1'/'2'), importe en pos 23-36, fechas DDMMYY
    · Variante Unicaja:  D/C en pos 22 ('0'='haber','1'='debe'), importe en pos 28-41,
                         fechas YYMMDD, conceptos en registros tipo '23'
  - CSV genérico (.csv) con cabeceras flexibles

La detección de variante es automática: si alguna línea de movimiento tiene '0'
en pos 22 con 14 dígitos en pos 28-41, se activa el parser Unicaja.
"""

import csv
import io
import re
import uuid
from datetime import datetime


# ── TIPOS POR CÓDIGO AEB ──────────────────────────────────────────────────────

_CODIGO_AEB_TIPO = {
    "01": "TRANSFERENCIA",   # Transferencia recibida
    "02": "TRANSFERENCIA",   # Transferencia emitida
    "03": "DOMICILIACION",   # Domiciliación cobrada
    "04": "DOMICILIACION",   # Domiciliación pagada
    "05": "OTROS",           # Cheque emitido
    "06": "OTROS",           # Cheque recibido
    "07": "TARJETA",         # Tarjeta
    "12": "OTROS",           # Intereses recibidos
    "13": "OTROS",           # Intereses pagados
    "14": "COMISION",        # Comisiones y gastos bancarios
    "50": "OTROS",           # Impuestos / retenciones
    "99": "OTROS",           # Otros
}

_TIPO_PALABRAS = {
    "COMISION":      ["COMISION", "COMISIÓN", "MANTENIMIENTO CTA", "GASTOS ADMINIST", "CUOTA"],
    "DOMICILIACION": ["DOMICILI", "RECIBO ", "ENDESA", "IBERDROLA", "TELEFON", "AGUA ", "GAS ", "SUMINISTRO"],
    "TARJETA":       ["TARJETA", "VISA", "MASTERCARD", "TPV"],
    "TRANSFERENCIA": ["TRANSFERENCIA", "TRANSF.", "REMESA", "ORDEN PAGO", "TRF", "NOMINA", "NÓMINA"],
}


def _detectar_tipo_texto(concepto: str) -> str:
    upper = concepto.upper()
    for tipo, palabras in _TIPO_PALABRAS.items():
        if any(p in upper for p in palabras):
            return tipo
    return "OTROS"


def _tipo_desde_codigo(codigo: str, concepto: str) -> str:
    return _CODIGO_AEB_TIPO.get(codigo.strip(), _detectar_tipo_texto(concepto))


# ── PARSEO DE FECHAS ──────────────────────────────────────────────────────────

def _fecha_ddmmyy(raw: str) -> str:
    """DDMMYY (6 dígitos) → YYYY-MM-DD. Estándar AEB."""
    try:
        return datetime.strptime(raw, "%d%m%y").strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return datetime.now().strftime("%Y-%m-%d")


def _fecha_yymmdd(raw: str) -> str:
    """YYMMDD (6 dígitos) → YYYY-MM-DD. Unicaja."""
    try:
        return datetime.strptime(raw, "%y%m%d").strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return datetime.now().strftime("%Y-%m-%d")


def _fecha_n43(raw: str) -> str:
    """DDMMYYYY (8 dígitos) → YYYY-MM-DD. Variante dinámica heredada."""
    try:
        return datetime.strptime(raw, "%d%m%Y").strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return datetime.now().strftime("%Y-%m-%d")


def _importe_n43(raw: str, dc: str) -> float:
    """14 dígitos con 2 decimales implícitos. dc='2' = cargo (negativo)."""
    try:
        valor = int(raw) / 100.0
        return round(-valor if dc == "2" else valor, 2)
    except (ValueError, TypeError):
        return 0.0


# ── DETECCIÓN DE VARIANTE N43 ─────────────────────────────────────────────────

def _detectar_variante_n43(lineas: list[str]) -> str:
    """
    Detecta la variante del fichero N43:
      'UNICAJA'  → pos 22 es '0' con 14 dígitos en pos 28-41 (Unicaja/CSB no estándar)
      'ESTANDAR' → pos 22 es '1'/'2' con 14 dígitos en pos 23-36 (AEB estándar)
      'DINAMICO' → usar _buscar_dc_pos (variantes con fechas DDMMYYYY tras D/C)
    """
    for linea in lineas:
        if len(linea) < 42 or linea[:2] != "22":
            continue
        dc = linea[22]
        # Variante Unicaja: usa '0' para haber (nunca aparece en estándar)
        if dc == '0' and linea[28:42].isdigit():
            return 'UNICAJA'
        # Variante estándar: '1' o '2' con importe en pos 23-36
        if dc in ('1', '2') and linea[23:37].isdigit():
            return 'ESTANDAR'
    return 'DINAMICO'


# ── PARSER DINÁMICO (variante heredada) ───────────────────────────────────────

def _buscar_dc_pos(line: str) -> int | None:
    """
    Busca la posición del indicador D/C en variantes N43 que sitúan
    D/C seguido de dos fechas DDMMYYYY y luego 14 dígitos de importe.
    """
    for i in range(20, 32):
        if i + 31 > len(line):
            break
        if line[i] not in ("1", "2"):
            continue
        if not (line[i + 1:i + 17].isdigit() and line[i + 17:i + 31].isdigit()):
            continue
        try:
            d1 = datetime.strptime(line[i + 1:i + 9], "%d%m%Y")
            d2 = datetime.strptime(line[i + 9:i + 17], "%d%m%Y")
            if not (2000 <= d1.year <= 2099 and 2000 <= d2.year <= 2099):
                continue
            return i
        except ValueError:
            continue
    return None


# ── PARSER NORMA 43 ────────────────────────────────────────────────────────────

def parsear_norma43(contenido: str, id_banco: int, id_usuario: str) -> list[dict]:
    """
    Parsea fichero AEB Norma 43 / CSB43.

    Soporta tres variantes:
      UNICAJA  — D/C en pos 22 ('0'=haber, '1'=debe, '9'=haber especial),
                 importe en pos 28-41, fechas YYMMDD, conceptos en registros 23xx.
      ESTANDAR — D/C en pos 22 ('1'=haber, '2'=debe),
                 importe en pos 23-36, fechas DDMMYY, conceptos en registros 22 sin D/C.
      DINAMICO — búsqueda dinámica del D/C (variante con fechas DDMMYYYY tras D/C).
    """
    lineas = contenido.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    variante = _detectar_variante_n43(lineas)

    movimientos: list[dict] = []
    mov_actual: dict | None = None
    conceptos_extra: list[str] = []

    def _guardar():
        nonlocal mov_actual, conceptos_extra
        if mov_actual is None:
            return
        if conceptos_extra:
            extra = " ".join(c for c in conceptos_extra if c).strip()
            if extra:
                mov_actual["concepto"] = (mov_actual["concepto"] + " " + extra).strip()
        movimientos.append(mov_actual)
        mov_actual = None
        conceptos_extra = []

    def _mov_dict(fecha_op, fecha_val, importe, concepto, codigo_aeb):
        return {
            "id_banco":           id_banco,
            "fecha_operacion":    fecha_op,
            "fecha_valor":        fecha_val,
            "concepto":           concepto,
            "importe":            importe,
            "saldo_posterior":    None,
            "referencia_banco":   f"N43-{id_banco}-{fecha_op}-{uuid.uuid4().hex[:8].upper()}",
            "tipo":               _tipo_desde_codigo(codigo_aeb, concepto),
            "estado":             "PENDIENTE",
            "origen":             "FICHERO_N43",
            "id_usuario_importa": id_usuario,
        }

    for linea in lineas:
        linea = linea.rstrip("\n\r")
        if len(linea) < 2:
            continue

        tipo_reg = linea[:2]

        # ── Registros de movimiento principal ─────────────────────────────────
        if tipo_reg == "22":

            if variante == "UNICAJA":
                # ── Variante Unicaja ──────────────────────────────────────────
                # Estructura: tipo(2) + oficina(4) + cc_shared(2) + cc_own(2)
                #             + fecha_op YYMMDD(6) + fecha_val YYMMDD(6)   → pos 10-21
                #             + D/C extendido(1) + extra(5) → pos 22-27
                #             + importe(14)                → pos 28-41
                #             + referencias y concepto     → pos 42+
                if len(linea) < 42:
                    continue
                dc = linea[22]
                if dc not in ('0', '1', '9'):
                    continue
                if not linea[28:42].isdigit():
                    continue

                _guardar()

                fecha_op  = _fecha_yymmdd(linea[10:16])
                fecha_val = _fecha_yymmdd(linea[16:22])
                # Unicaja: '0' o '9' = HABER (positivo), '1' = DEBE (negativo)
                dc_std    = "2" if dc == "1" else "1"
                importe   = _importe_n43(linea[28:42], dc_std)
                codigo_aeb = linea[8:10]   # código propio del banco (pos 8-9)
                # El campo concepto en pos 42+ son referencias internas del banco
                # (números de documento, ref. SEPA, etc.). Se deja vacío y se
                # rellena con el texto real de las líneas de continuación 23xx.
                concepto  = ""

                mov_actual = _mov_dict(fecha_op, fecha_val, importe, concepto, codigo_aeb)
                conceptos_extra = []

            elif variante == "ESTANDAR":
                # ── Variante estándar AEB ─────────────────────────────────────
                # Estructura fija: D/C en pos 22, importe pos 23-36, fechas DDMMYY
                if len(linea) < 37:
                    continue
                dc = linea[22]
                if dc not in ('1', '2'):
                    # Sin D/C válido → línea de concepto adicional (estándar)
                    texto = linea[4:].strip()
                    if texto and mov_actual is not None:
                        conceptos_extra.append(texto)
                    continue
                if not linea[23:37].isdigit():
                    continue

                _guardar()

                fecha_op  = _fecha_ddmmyy(linea[10:16])
                fecha_val = _fecha_ddmmyy(linea[16:22])
                importe   = _importe_n43(linea[23:37], dc)
                codigo_aeb = linea[6:8]    # código compartido AEB (pos 6-7)
                concepto  = linea[37:].strip()

                mov_actual = _mov_dict(fecha_op, fecha_val, importe, concepto, codigo_aeb)
                conceptos_extra = []

            else:
                # ── Variante dinámica (búsqueda de D/C) ──────────────────────
                dc_pos = _buscar_dc_pos(linea)
                if dc_pos is not None:
                    _guardar()
                    dc        = linea[dc_pos]
                    fecha_op  = _fecha_n43(linea[dc_pos + 1:dc_pos + 9])
                    fecha_val = _fecha_n43(linea[dc_pos + 9:dc_pos + 17])
                    importe   = _importe_n43(linea[dc_pos + 17:dc_pos + 31], dc)
                    concepto  = linea[dc_pos + 31:].strip()
                    codigo_aeb = linea[dc_pos - 2:dc_pos].strip() if dc_pos >= 2 else ""
                    mov_actual = _mov_dict(fecha_op, fecha_val, importe, concepto, codigo_aeb)
                    conceptos_extra = []
                else:
                    texto = linea[4:].strip()
                    if texto and mov_actual is not None:
                        conceptos_extra.append(texto)

        # ── Registros de concepto adicional (tipo 23xx) ───────────────────────
        elif tipo_reg == "23":
            # Unicaja y otras variantes: líneas de concepto empiezan con "23"
            # Estructura: "23" + sub-código (2 chars) + texto (76 chars)
            texto = linea[4:].strip()
            if texto and mov_actual is not None:
                conceptos_extra.append(texto)

        # ── Registros de cierre ───────────────────────────────────────────────
        elif tipo_reg in ("33", "88", "99"):
            _guardar()

    _guardar()
    return movimientos


# ── PARSER CSV GENÉRICO ────────────────────────────────────────────────────────

_CSV_ALIASES: dict[str, list[str]] = {
    "fecha":    ["fecha", "fecha operacion", "fecha_operacion", "date",
                 "f. operación", "f.operacion", "fecha op"],
    "concepto": ["concepto", "descripcion", "descripción", "concept",
                 "motivo", "detalle", "observaciones"],
    "importe":  ["importe", "amount", "importe (eur)", "cargo/abono",
                 "importe eur", "cargo", "abono"],
    "saldo":    ["saldo", "balance", "saldo (eur)", "saldo posterior", "saldo eur"],
}


def _mapear_cabeceras(cabeceras: list[str]) -> dict[str, int]:
    mapa: dict[str, int] = {}
    for campo, aliases in _CSV_ALIASES.items():
        for i, cab in enumerate(cabeceras):
            if cab.lower().strip() in aliases:
                mapa[campo] = i
                break
    return mapa


def _parsear_fecha_csv(raw: str) -> str | None:
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parsear_importe_csv(raw: str) -> float:
    """Maneja tanto formato español (1.234,56) como anglosajón (1,234.56)."""
    s = raw.strip().replace(" ", "")
    if re.search(r"\d\.\d{3},", s):          # 1.234,56 → español
        s = s.replace(".", "").replace(",", ".")
    elif re.search(r"\d,\d{3}\.", s):         # 1,234.56 → anglosajón
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")               # 1234,56 simple
    return float(s)


def parsear_csv(contenido: str, id_banco: int, id_usuario: str) -> list[dict]:
    sep = ";" if contenido.count(";") > contenido.count(",") else ","
    reader = csv.reader(io.StringIO(contenido), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []

    mapa = _mapear_cabeceras(rows[0])
    if "fecha" not in mapa or "importe" not in mapa:
        return []

    movimientos: list[dict] = []
    for row in rows[1:]:
        if not any(cell.strip() for cell in row):
            continue
        try:
            fecha = _parsear_fecha_csv(row[mapa["fecha"]])
            if not fecha:
                continue
            importe = _parsear_importe_csv(row[mapa["importe"]])
            concepto = row[mapa["concepto"]].strip() if "concepto" in mapa else ""
            saldo = None
            if "saldo" in mapa:
                try:
                    saldo = _parsear_importe_csv(row[mapa["saldo"]])
                except (ValueError, IndexError):
                    pass
            movimientos.append({
                "id_banco":         id_banco,
                "fecha_operacion":  fecha,
                "fecha_valor":      fecha,
                "concepto":         concepto,
                "importe":          round(importe, 2),
                "saldo_posterior":  saldo,
                "referencia_banco": f"CSV-{id_banco}-{fecha}-{uuid.uuid4().hex[:8].upper()}",
                "tipo":             _detectar_tipo_texto(concepto),
                "estado":           "PENDIENTE",
                "origen":           "FICHERO_CSV",
                "id_usuario_importa": id_usuario,
            })
        except (ValueError, IndexError):
            continue

    return movimientos


# ── DETECCIÓN AUTOMÁTICA ──────────────────────────────────────────────────────

_OLE2_MAGIC = b"\xd0\xcf\x11\xe0"   # cabecera fichero Excel binario (.xls)


def es_excel_binario(contenido_bytes: bytes) -> bool:
    return contenido_bytes[:4] == _OLE2_MAGIC


def _fecha_excel_serial(serial) -> str:
    """Convierte número serial de Excel (float) a cadena YYYY-MM-DD."""
    from datetime import timedelta
    try:
        d = datetime(1899, 12, 30) + timedelta(days=int(float(serial)))
        return d.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return datetime.now().strftime("%Y-%m-%d")


def _es_xls_unicaja_web(ws) -> bool:
    """Detecta el formato tabular de exportación web de Unicaja (cabeceras en fila 10)."""
    if ws.nrows < 11:
        return False
    primera = str(ws.cell_value(10, 0)).strip().lower()
    return "fecha" in primera and "operaci" in primera


def parsear_xls_unicaja_web(ws, id_banco: int, id_usuario: str) -> list[dict]:
    """
    Parsea el Excel tabular que genera Unicaja desde su banca online:
      Fila 10 → cabeceras (Fecha de operación, Fecha valor, Concepto, Importe, Divisa, Saldo…)
      Fila 11+ → datos
    Fechas como seriales de Excel (float), importe positivo=abono/negativo=cargo.
    """
    movimientos: list[dict] = []
    for row_idx in range(11, ws.nrows):
        try:
            fecha_op_raw  = ws.cell_value(row_idx, 0)
            fecha_val_raw = ws.cell_value(row_idx, 1)
            concepto      = str(ws.cell_value(row_idx, 2)).strip()
            importe_raw   = ws.cell_value(row_idx, 3)
            saldo_raw     = ws.cell_value(row_idx, 5)

            if not fecha_op_raw or not concepto:
                continue

            fecha_op  = _fecha_excel_serial(fecha_op_raw)
            fecha_val = _fecha_excel_serial(fecha_val_raw) if fecha_val_raw else fecha_op
            importe   = round(float(importe_raw), 2)
            saldo     = round(float(saldo_raw), 2) if saldo_raw != "" else None

            movimientos.append({
                "id_banco":            id_banco,
                "fecha_operacion":     fecha_op,
                "fecha_valor":         fecha_val,
                "concepto":            concepto,
                "importe":             importe,
                "saldo_posterior":     saldo,
                "referencia_banco":    f"UNICAJA-{id_banco}-{fecha_op}-{uuid.uuid4().hex[:8].upper()}",
                "tipo":                _detectar_tipo_texto(concepto),
                "estado":              "PENDIENTE",
                "origen":              "FICHERO_XLS",
                "id_usuario_importa":  id_usuario,
            })
        except (ValueError, TypeError, IndexError):
            continue

    return movimientos


def parsear_xls_n43(contenido_bytes: bytes, id_banco: int, id_usuario: str) -> list[dict]:
    """
    Lee un .xls de Unicaja:
      - Si es el formato tabular web (cabeceras en fila 10) → parsear_xls_unicaja_web
      - Si contiene líneas N43 embebidas en celdas → parsear_norma43
    """
    import xlrd
    wb = xlrd.open_workbook(file_contents=contenido_bytes)
    ws = wb.sheet_by_index(0)

    # Formato tabular web de Unicaja
    if _es_xls_unicaja_web(ws):
        return parsear_xls_unicaja_web(ws, id_banco, id_usuario)

    # Intento de N43 embebido en celdas
    n43_lines = []
    for row_idx in range(ws.nrows):
        for col_idx in range(ws.ncols):
            val = str(ws.cell_value(row_idx, col_idx)).strip()
            if val[:2] in ("11", "22", "23", "33", "88", "99") and len(val) >= 2:
                n43_lines.append(val)
                break

    if not n43_lines or not any(l.startswith("11") for l in n43_lines):
        raise ValueError(
            "El fichero Excel no contiene un extracto N43 ni un formato tabular reconocido. "
            "Descarga el extracto desde Unicaja: Mis cuentas → Movimientos → Exportar → "
            "selecciona AEB, Norma 43 o el formato Excel de movimientos."
        )

    return parsear_norma43("\n".join(n43_lines), id_banco, id_usuario)


def parsear_xlsx_santander(contenido_bytes: bytes, id_banco: int, id_usuario: str) -> list[dict]:
    """
    Parsea el xlsx que genera Santander desde su banca online:
      Fila 8 → cabeceras (Fecha Operación, Fecha Valor, Concepto, Importe, Divisa, Saldo…)
      Fila 9+ → datos; fechas en dd/mm/yyyy, importes float (+abono/-cargo).
    """
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(contenido_bytes), data_only=True)
    ws = wb.active

    movimientos: list[dict] = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        try:
            fecha_op_raw, fecha_val_raw, concepto, importe_raw, _, saldo_raw = row[:6]
            if not fecha_op_raw or not concepto or importe_raw is None:
                continue

            fecha_op  = _parsear_fecha_csv(str(fecha_op_raw))
            fecha_val = _parsear_fecha_csv(str(fecha_val_raw)) if fecha_val_raw else fecha_op
            if not fecha_op:
                continue
            importe = round(float(importe_raw), 2)
            saldo   = round(float(saldo_raw), 2) if saldo_raw is not None else None

            movimientos.append({
                "id_banco":           id_banco,
                "fecha_operacion":    fecha_op,
                "fecha_valor":        fecha_val or fecha_op,
                "concepto":           str(concepto).strip(),
                "importe":            importe,
                "saldo_posterior":    saldo,
                "referencia_banco":   f"SANTANDER-{id_banco}-{fecha_op}-{uuid.uuid4().hex[:8].upper()}",
                "tipo":               _detectar_tipo_texto(str(concepto)),
                "estado":             "PENDIENTE",
                "origen":             "FICHERO_XLSX",
                "id_usuario_importa": id_usuario,
            })
        except (ValueError, TypeError):
            continue

    return movimientos


def _es_xlsx_santander(contenido_bytes: bytes) -> bool:
    """Detecta el xlsx tabular de Santander comprobando la cabecera de la fila 8."""
    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(contenido_bytes), data_only=True, read_only=True)
        ws = next(iter(wb))
        for i, row in enumerate(ws.iter_rows(min_row=8, max_row=8, values_only=True), 1):
            primera = str(row[0] or "").lower()
            return "fecha" in primera and "operaci" in primera
    except Exception:
        pass
    return False


def detectar_y_parsear(
    contenido: str,
    nombre_fichero: str,
    id_banco: int,
    id_usuario: str,
    contenido_bytes: bytes | None = None,
) -> list[dict]:
    """
    Detecta el formato y aplica el parser correspondiente:
      - .xlsx Santander (cabecera en fila 8) → parsear_xlsx_santander
      - Excel binario .xls → parsear_xls_n43 (Unicaja web o N43 embebido)
      - Extensión n43/43/sta/csb o cabecera "11" → Norma 43 texto
      - Resto → CSV genérico
    """
    ext = nombre_fichero.lower().rsplit(".", 1)[-1] if "." in nombre_fichero else ""

    if ext == "xlsx" and contenido_bytes:
        if _es_xlsx_santander(contenido_bytes):
            return parsear_xlsx_santander(contenido_bytes, id_banco, id_usuario)

    if contenido_bytes and es_excel_binario(contenido_bytes):
        return parsear_xls_n43(contenido_bytes, id_banco, id_usuario)

    primera_linea = contenido.lstrip("﻿").strip()[:10]

    if ext in ("n43", "43", "sta", "csb", "q43") or primera_linea.startswith("11"):
        return parsear_norma43(contenido.lstrip("﻿"), id_banco, id_usuario)

    return parsear_csv(contenido, id_banco, id_usuario)
